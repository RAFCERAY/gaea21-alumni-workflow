"""Utilitaires pour le workflow d'enrichissement alumni Gaea21.

Ce module contient :
- Classification automatique du pays depuis la ville
- Classification automatique du secteur depuis l'entreprise
- Extraction d'infos depuis le texte LinkedIn
- Gestion du fichier Excel INTERNE
"""

import re
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

# ============================================================================
# DICTIONNAIRES DE CLASSIFICATION
# ============================================================================

# Mapping ville → pays (enrichi au fur et à mesure)
CITY_TO_COUNTRY = {
    # Suisse
    "genève": "Suisse", "geneva": "Suisse", "lausanne": "Suisse",
    "zurich": "Suisse", "bern": "Suisse", "berne": "Suisse",
    "basel": "Suisse", "bâle": "Suisse", "sierre": "Suisse",
    "fribourg": "Suisse", "estavayer": "Suisse", "payerne": "Suisse",
    "vaud": "Suisse", "valais": "Suisse", "bogis-bossey": "Suisse",
    "mies": "Suisse", "gex": "France",
    # France
    "paris": "France", "lyon": "France", "marseille": "France",
    "toulouse": "France", "bordeaux": "France", "lille": "France",
    "rennes": "France", "nantes": "France", "strasbourg": "France",
    "montpellier": "France", "grenoble": "France", "annecy": "France",
    "chamonix": "France", "pau": "France", "nice": "France",
    "evry": "France", "orléans": "France", "rouen": "France",
    "dijon": "France", "martinique": "France",
    # Allemagne
    "berlin": "Allemagne", "munich": "Allemagne", "münchen": "Allemagne",
    "hamburg": "Allemagne", "stuttgart": "Allemagne", "frankfurt": "Allemagne",
    # Maroc
    "casablanca": "Maroc", "rabat": "Maroc", "marrakech": "Maroc",
    "fes": "Maroc", "fès": "Maroc", "tanger": "Maroc",
    "khouribga": "Maroc", "agadir": "Maroc",
    # Tunisie
    "tunis": "Tunisie", "sfax": "Tunisie", "nabeul": "Tunisie",
    "sousse": "Tunisie",
    # Belgique
    "bruxelles": "Belgique", "brussels": "Belgique", "anvers": "Belgique",
    "antwerp": "Belgique", "liège": "Belgique",
    # Royaume-Uni
    "london": "Royaume-Uni", "londres": "Royaume-Uni",
    "manchester": "Royaume-Uni", "edinburgh": "Royaume-Uni",
    # Espagne
    "madrid": "Espagne", "barcelona": "Espagne", "barcelone": "Espagne",
    "valencia": "Espagne", "sevilla": "Espagne",
    # Italie
    "roma": "Italie", "rome": "Italie", "milano": "Italie", "milan": "Italie",
    # Luxembourg
    "luxembourg": "Luxembourg",
    # Portugal
    "lisbon": "Portugal", "lisbonne": "Portugal", "porto": "Portugal",
    # USA/Canada
    "new york": "États-Unis", "san francisco": "États-Unis",
    "washington": "États-Unis", "boston": "États-Unis", "chicago": "États-Unis",
    "los angeles": "États-Unis",
    "montréal": "Canada", "montreal": "Canada", "toronto": "Canada",
    "vancouver": "Canada",
    # Afrique
    "dakar": "Sénégal", "abidjan": "Côte d'Ivoire", "ouagadougou": "Burkina Faso",
    "lomé": "Togo", "lome": "Togo", "cotonou": "Bénin",
    "alger": "Algérie", "oran": "Algérie",
    # Amérique Latine
    "méxico": "Mexique", "mexico": "Mexique", "cdmx": "Mexique",
    "são paulo": "Brésil", "sao paulo": "Brésil", "rio": "Brésil",
}

# Indicateurs de pays dans le texte (fallback si ville non reconnue)
COUNTRY_KEYWORDS = {
    "Suisse": ["suisse", "switzerland", "helvetia", "schweiz"],
    "France": ["france", "île-de-france", "auvergne-rhône-alpes",
               "provence-alpes", "hauts-de-france", "bretagne",
               "nouvelle-aquitaine", "grand est", "pays de la loire"],
    "Allemagne": ["germany", "allemagne", "deutschland", "baden-württemberg"],
    "Maroc": ["morocco", "maroc", "rabat-salé", "béni mellal"],
    "Tunisie": ["tunisia", "tunisie", "gouvernorat"],
    "Belgique": ["belgique", "belgium", "belgië", "wallonie", "flandres"],
    "Royaume-Uni": ["united kingdom", "royaume-uni", "england", "uk"],
    "Luxembourg": ["luxembourg"],
    "Mexique": ["mexico", "mexique", "mexican"],
    "Canada": ["canada", "quebec", "québec", "ontario"],
    "Brésil": ["brazil", "brasil", "brésil"],
    "Kenya": ["kenya", "nairobi"],
    "Sénégal": ["senegal", "sénégal"],
    "Togo": ["togo", "lome"],
    "Côte d'Ivoire": ["ivory coast", "côte d'ivoire"],
    "Espagne": ["spain", "españa", "espagne"],
    "Portugal": ["portugal"],
    "Italie": ["italy", "italia", "italie"],
    "États-Unis": ["united states", "usa", "america", "american"],
}

# Mots-clés d'entreprises/domaines → secteurs
SECTOR_KEYWORDS = {
    "Environnement / Climat": [
        "sustainability", "sustainable", "durabilité", "durable",
        "climate", "climat", "carbon", "carbone", "ges", "greenhouse",
        "environnement", "environmental", "écologie", "ecology",
        "biodiversité", "biodiversity", "renewable", "renouvelable",
        "circular", "circulaire", "rse", "esg", "csr", "lca", "acv",
        "agroecology", "agroécologie", "agricultural", "agronomie",
        "wwf", "greenpeace", "patagonia", "ellen macarthur",
        "bilan carbone", "décarbonation", "decarbonation",
        "vertical farming", "ecovadis", "sbti", "iso 14001",
    ],
    "Finance durable / ESG": [
        "finance durable", "sustainable finance", "impact investing",
        "esg", "mirova", "triodos", "sustainable banking",
        "green bonds", "bank al-maghrib", "banque centrale",
        "responsible investment",
    ],
    "Tech / Data": [
        "data", "scientist", "analyst", "analytics", "python", "sql",
        "machine learning", "ml", "ai", "artificial intelligence",
        "software", "developer", "développeur", "fullstack",
        "react", "angular", "node", "devops", "cloud", "aws", "azure",
        "cybersecurity", "cybersécurité", "security", "fullstack",
        "big data", "data engineer", "it akademy", "ensiasd", "insea",
        "développement web", "programmation", "backend", "frontend",
    ],
    "Conseil / Audit": [
        "consulting", "consultant", "advisory", "conseil", "auditor",
        "auditeur", "mckinsey", "bcg", "bain", "deloitte", "pwc",
        "ey", "kpmg", "accenture", "strategy", "cfo-as-a-service",
        "recruteur", "recrutement", "coach professionnel",
    ],
    "Éducation / Recherche": [
        "université", "university", "école", "school", "professor",
        "professeur", "teacher", "researcher", "chercheur", "phd",
        "doctorat", "research", "recherche", "étudiant", "student",
        "master", "formation", "formateur", "formatrice", "teaching",
        "stagiaire ingénieur", "skema", "hohenheim", "insea",
        "ensiasd", "university of", "institut", "academy", "akademy",
    ],
    "ONG / Associatif": [
        "ngo", "ong", "nonprofit", "non-profit", "association",
        "charity", "fondation", "foundation", "volunteer", "bénévole",
        "croix-rouge", "red cross", "amnesty", "oxfam", "unicef",
        "médecins sans frontières",
    ],
    "Secteur public": [
        "government", "gouvernement", "ministry", "ministère",
        "onu", "un.org", "united nations", "union européenne",
        "european union", "commission européenne", "ministère",
        "fonctionnaire", "service public", "administration",
        "préfecture", "mairie", "région",
    ],
    "Santé": [
        "health", "santé", "hospital", "hôpital", "medical", "médical",
        "nurse", "infirmière", "doctor", "médecin", "pharma",
        "pharmaceutical", "biotech", "coach eft", "coaching bien-être",
        "thérapeute", "psychologue",
    ],
    "Entreprise privée": [
        "sa", "sas", "sarl", "gmbh", "ltd", "inc", "corp",
    ],
}

# Entreprises connues → secteur (dictionnaire manuel enrichi)
KNOWN_COMPANIES = {
    "honda": "Entreprise privée",
    "nestlé": "Entreprise privée",
    "nestle": "Entreprise privée",
    "patagonia": "Environnement / Climat",
    "wwf": "ONG / Associatif",
    "orange": "Entreprise privée",
    "css": "Entreprise privée",
    "numtech": "Environnement / Climat",
    "wesh grow": "Environnement / Climat",
    "pronoia": "Entreprise privée",
    "kookai": "Entreprise privée",
    "it akademy": "Éducation / Recherche",
    "skema": "Éducation / Recherche",
    "skema business school": "Éducation / Recherche",
    "university of hohenheim": "Éducation / Recherche",
    "université de fribourg": "Éducation / Recherche",
    "ensiasd": "Éducation / Recherche",
    "insea": "Éducation / Recherche",
    "imt atlantique": "Éducation / Recherche",
    "bank al-maghrib": "Finance durable / ESG",
    "honda motor europe": "Entreprise privée",
    "gaea21": "ONG / Associatif",
    "gaea": "ONG / Associatif",
    "rayen academy": "Éducation / Recherche",
    "groupe partouche": "Entreprise privée",
    "moho consulting": "Conseil / Audit",
    "altaservice": "Conseil / Audit",
    "novasolvd": "Conseil / Audit",
}


# ============================================================================
# FONCTIONS DE CLASSIFICATION
# ============================================================================

def clean_text(text: str) -> str:
    """Nettoie et normalise un texte pour la classification."""
    if not text or pd.isna(text):
        return ""
    return str(text).lower().strip()


def guess_country_from_location(location: str) -> Optional[str]:
    """Devine le pays à partir d'une localisation."""
    if not location:
        return None
    loc = clean_text(location)

    # 1. Recherche exacte dans le dictionnaire villes
    for city, country in CITY_TO_COUNTRY.items():
        if city in loc:
            return country

    # 2. Recherche par mots-clés pays
    for country, keywords in COUNTRY_KEYWORDS.items():
        for kw in keywords:
            if kw in loc:
                return country

    return None


def guess_sector_from_company(company: str, headline: str = "") -> Optional[str]:
    """Devine le secteur à partir de l'entreprise et du headline."""
    combined = clean_text(company) + " " + clean_text(headline)
    if not combined.strip():
        return None

    # 1. Entreprises connues
    for company_kw, sector in KNOWN_COMPANIES.items():
        if company_kw in combined:
            return sector

    # 2. Mots-clés par secteur (on prend le secteur avec le plus de matchs)
    scores = {}
    for sector, keywords in SECTOR_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw in combined)
        if score > 0:
            scores[sector] = score

    if scores:
        return max(scores, key=scores.get)

    return None


def extract_info_from_linkedin_text(text: str) -> dict:
    """Extrait les informations d'un texte LinkedIn collé.

    Le texte LinkedIn contient typiquement :
    - Ligne 1 : Prénom Nom
    - Ligne 2 : relation (ignorée)
    - Ligne 3 : Headline (poste actuel)
    - Ligne 4-5 : Entreprise + école
    - Ligne suivante : Localisation
    """
    if not text:
        return {}

    lines = [l.strip() for l in text.split('\n') if l.strip()]
    result = {
        "nom_complet": "",
        "headline": "",
        "entreprise": "",
        "localisation": "",
        "poste": "",
        "pays_devine": "",
        "secteur_devine": "",
        "ville_devinee": "",
    }

    # Skip lignes "relation de X niveau"
    filtered = [l for l in lines if not re.match(r"relation de \d+", l, re.I)
                and not re.match(r"\d+e\s*$", l, re.I)]

    if not filtered:
        return result

    # Heuristique : 1er élément = nom, 2e = headline
    if len(filtered) >= 1:
        result["nom_complet"] = filtered[0]
    if len(filtered) >= 2:
        result["headline"] = filtered[1]
        result["poste"] = filtered[1]

    # Localisation = ligne contenant "Coordonnées" ou juste avant
    for i, line in enumerate(filtered):
        if "coordonnées" in line.lower():
            if i > 0:
                result["localisation"] = filtered[i - 1]
            break

    # Si pas trouvé, chercher des patterns géographiques
    if not result["localisation"]:
        for line in filtered[2:6]:  # entre ligne 3 et 6
            if any(kw in clean_text(line) for kw in
                   ["france", "suisse", "maroc", "tunisie", "allemagne",
                    "genève", "paris", "berlin", "london"]):
                result["localisation"] = line
                break

    # Entreprise = extraire du headline si format "Poste chez/at Entreprise"
    if result["headline"]:
        h = result["headline"]
        patterns = [
            r"(?:chez|at|@)\s+([^|•]+)",
            r"\|\s*([A-Z][^|]+?)(?:\s*$|\s*\|)",
        ]
        for pattern in patterns:
            match = re.search(pattern, h)
            if match:
                result["entreprise"] = match.group(1).strip()
                break

    # Devine pays
    result["pays_devine"] = guess_country_from_location(result["localisation"]) or ""

    # Devine secteur
    result["secteur_devine"] = guess_sector_from_company(
        result["entreprise"], result["headline"]
    ) or ""

    return result


# ============================================================================
# GESTION DU FICHIER EXCEL
# ============================================================================

# Mapping colonnes fichier INTERNE (1-indexed Excel)
COLS_INTERNE = {
    "id": 1,
    "prenom": 2,
    "nom": 3,
    "annee_sortie": 4,
    "fonction_gaea": 5,
    "contrat": 6,
    "date_entree": 7,
    "date_sortie": 8,
    "email_pro": 9,
    "email_perso": 10,
    "motif": 11,
    "statut_alumni": 12,
    "flagged": 13,
    "separator": 14,
    "poste_actuel": 15,
    "entreprise": 16,
    "ville": 17,
    "pays": 18,
    "secteur": 19,
    "url_linkedin": 20,
    "statut_enrich": 21,
    "notes": 22,
}


def load_alumni_from_excel(file_path: str, sheet_name: str = None) -> pd.DataFrame:
    """Charge le fichier Excel dans un DataFrame pandas.

    Auto-détecte la bonne feuille (Alumni AG, Alumni (tous), etc.)
    et la bonne ligne d'en-tête (0, 1 ou 2).
    """
    # Liste des feuilles candidates, dans l'ordre de priorité
    candidate_sheets = [
        "Alumni AG", "Alumni (tous)", "Alumni tous", "Alumni",
        "Sorties", "Alumnis",
    ]

    # Si sheet_name donné, on le met en premier
    if sheet_name:
        candidate_sheets.insert(0, sheet_name)

    # Récupère les feuilles existantes dans le fichier
    xl = pd.ExcelFile(file_path)
    existing_sheets = xl.sheet_names

    # Trouve la 1ère feuille qui existe parmi les candidates
    chosen_sheet = None
    for cand in candidate_sheets:
        if cand in existing_sheets:
            chosen_sheet = cand
            break

    if chosen_sheet is None:
        # Fallback : première feuille du fichier
        chosen_sheet = existing_sheets[0]

    # Essaie plusieurs positions d'en-tête (0, 1, 2)
    for header_row in [1, 0, 2]:
        try:
            df = pd.read_excel(file_path, sheet_name=chosen_sheet, header=header_row)
            df.columns = [str(c).strip() for c in df.columns]
            # Valide : doit contenir au moins "Nom" ou "Prénom"
            if any("Nom" in c or "Prénom" in c or "Prenom" in c for c in df.columns):
                return df
        except Exception:
            continue

    # Dernier fallback
    df = pd.read_excel(file_path, sheet_name=chosen_sheet, header=1)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _find_status_column(df: pd.DataFrame) -> str:
    """Trouve le nom de colonne Statut, peu importe la variante."""
    candidates = ["Statut enrich.", "Statut", "Statut enrichissement",
                  "Statut enrich", "Statut_enrich"]
    for c in candidates:
        if c in df.columns:
            return c
    # Fallback : première colonne contenant 'statut'
    for c in df.columns:
        if "statut" in str(c).lower():
            return c
    return "Statut"  # défaut


def get_next_alumni_to_enrich(
    df: pd.DataFrame,
    priority_years: list = None,
    priority_functions: list = None
) -> Optional[dict]:
    """Retourne le prochain alumni à enrichir, par ordre de priorité.

    Priorité : année la plus récente > fonctions vedettes > ordre alphabétique
    """
    if priority_years is None:
        priority_years = [2024, 2025, 2023, 2026]

    status_col = _find_status_column(df)

    # Filtre les non-enrichis en utilisant les emojis (plus fiable que les mots)
    # ✅ = Enrichi, ❌ = Introuvable, 🟡 = Privé
    statut_str = df[status_col].astype(str).str.lower()

    def is_traite(val):
        """Détecte si un alumni a déjà été traité."""
        if pd.isna(val) or val in ("", "nan", "none"):
            return False
        v = str(val).lower().strip()
        # On cherche l'emoji ou le mot spécifique
        has_check = "✅" in v or "enrichi" in v
        has_cross = "❌" in v or "introuvable" in v
        has_yellow = "🟡" in v or "privé" in v or "prive" in v
        # Mais on EXCLUT "à enrichir" qui contient "enrichi"
        is_placeholder = "à enrich" in v or "a enrich" in v or "⬜" in v
        if is_placeholder:
            return False
        return has_check or has_cross or has_yellow

    mask_traite = df[status_col].apply(is_traite)
    to_enrich = df[~mask_traite].copy()

    if len(to_enrich) == 0:
        return None

    # Priorité par année
    annee_col = "Année sortie" if "Année sortie" in df.columns else "Annee sortie"
    to_enrich["priority_year"] = to_enrich[annee_col].apply(
        lambda y: priority_years.index(int(y)) if (pd.notna(y) and str(y).replace(".0", "").isdigit() and int(float(y)) in priority_years) else 999
    )

    # Trie
    sort_cols = ["priority_year"]
    if "Nom" in df.columns:
        sort_cols.append("Nom")
    if "Prénom" in df.columns:
        sort_cols.append("Prénom")
    to_enrich = to_enrich.sort_values(sort_cols)

    # Retourne le premier
    row = to_enrich.iloc[0]
    return {
        "row_index_excel": row.name + 3,  # +3 car Excel 1-indexed + 2 lignes d'en-tête
        "id": row.get("ID", ""),
        "prenom": row.get("Prénom", ""),
        "nom": row.get("Nom", ""),
        "annee_sortie": row.get(annee_col, ""),
        "fonction_gaea": row.get("Fonction Gaea21", ""),
        "date_entree": row.get("Date entrée", ""),
        "date_sortie": row.get("Date sortie", ""),
    }


def save_enrichment_to_excel(
    file_path: str,
    row_index_excel: int,
    data: dict,
    sheet_name: str = "Alumni AG"
):
    """Sauvegarde les données enrichies dans la bonne ligne du fichier Excel.

    Auto-détecte les noms de colonnes (Statut vs Statut enrich.) et leurs positions.
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Trouve dynamiquement la ligne d'en-tête et les positions des colonnes
    # On cherche dans les 2 premières lignes
    header_row = None
    col_positions = {}

    for candidate_row in [1, 2, 3]:
        try:
            headers = [ws.cell(row=candidate_row, column=c).value for c in range(1, 30)]
        except Exception:
            continue
        headers_str = [str(h).strip() if h else "" for h in headers]

        # On valide que c'est bien la ligne d'en-tête si elle contient au moins "Prénom" ou "Nom"
        if any("Prénom" in h or "Prenom" in h or "Nom" in h for h in headers_str):
            header_row = candidate_row
            for idx, header in enumerate(headers_str, start=1):
                col_positions[header] = idx
            break

    if not col_positions:
        raise ValueError("Impossible de trouver la ligne d'en-tête dans le fichier Excel")

    # Mapping des clés data → noms de colonnes possibles
    col_aliases = {
        "Poste actuel": ["Poste actuel", "Poste"],
        "Entreprise": ["Entreprise", "Entreprise actuelle"],
        "Ville": ["Ville"],
        "Pays": ["Pays"],
        "Secteur": ["Secteur"],
        "URL LinkedIn": ["URL LinkedIn", "URL LinkedIn"],
        "Statut": ["Statut enrich.", "Statut", "Statut enrichissement"],
        "Notes": ["Notes"],
    }

    # Écrit chaque donnée dans la bonne colonne
    for data_key, value in data.items():
        if not value:
            continue
        # Trouve la position de la colonne
        aliases = col_aliases.get(data_key, [data_key])
        col_idx = None
        for alias in aliases:
            if alias in col_positions:
                col_idx = col_positions[alias]
                break
        if col_idx:
            ws.cell(row=row_index_excel, column=col_idx, value=value)

    wb.save(file_path)


def build_linkedin_search_url(prenom: str, nom: str, extra_keyword: str = "Gaea21") -> str:
    """Construit l'URL de recherche LinkedIn pré-remplie."""
    query = f"{prenom} {nom} {extra_keyword}".strip()
    query_encoded = query.replace(" ", "%20")
    return f"https://www.linkedin.com/search/results/people/?keywords={query_encoded}"


def _is_enrichi(val) -> bool:
    """Détecte si un alumni est enrichi (statut ✅ Enrichi)."""
    if pd.isna(val) or val in ("", "nan", "none"):
        return False
    v = str(val).lower().strip()
    if "à enrich" in v or "a enrich" in v or "⬜" in v:
        return False
    return "✅" in v or "enrichi" in v


def _is_introuvable(val) -> bool:
    """Détecte si un alumni est introuvable."""
    if pd.isna(val):
        return False
    v = str(val).lower().strip()
    return "❌" in v or "introuvable" in v


def _is_prive(val) -> bool:
    """Détecte si un alumni a un profil privé."""
    if pd.isna(val):
        return False
    v = str(val).lower().strip()
    return "🟡" in v or "privé" in v or "prive" in v


def get_stats(df: pd.DataFrame) -> dict:
    """Calcule les statistiques globales de la base."""
    status_col = _find_status_column(df)
    statut = df[status_col]
    total = len(df)
    enrichis = int(statut.apply(_is_enrichi).sum())
    introuvables = int(statut.apply(_is_introuvable).sum())
    prives = int(statut.apply(_is_prive).sum())

    return {
        "total": total,
        "enrichis": enrichis,
        "introuvables": introuvables,
        "prives": prives,
        "restants": int(total - enrichis - introuvables - prives),
        "pct_enrichis": round(100 * enrichis / total, 1) if total else 0,
    }


def get_top_entreprises(df: pd.DataFrame, n: int = 15) -> list:
    """Retourne les top N entreprises par nombre d'alumni."""
    status_col = _find_status_column(df)
    mask_enrichi = df[status_col].apply(_is_enrichi)
    entreprises = df[mask_enrichi]["Entreprise"]
    entreprises = entreprises.dropna().astype(str).str.strip()
    entreprises = entreprises[entreprises != ""]
    return entreprises.value_counts().head(n).to_dict()


def get_by_country(df: pd.DataFrame) -> dict:
    """Retourne la répartition par pays (alumni enrichis)."""
    status_col = _find_status_column(df)
    mask_enrichi = df[status_col].apply(_is_enrichi)
    pays = df[mask_enrichi]["Pays"]
    return pays.dropna().value_counts().to_dict()


def get_by_sector(df: pd.DataFrame) -> dict:
    """Retourne la répartition par secteur (alumni enrichis)."""
    status_col = _find_status_column(df)
    mask_enrichi = df[status_col].apply(_is_enrichi)
    sec = df[mask_enrichi]["Secteur"]
    return sec.dropna().value_counts().to_dict()
