"""Script one-shot : marque les 8 vedettes identifiées dans la colonne Notes.

Usage :
    python3 marquer_vedettes.py

Le script :
1. Ouvre alumni-gaea21-AG-READY.xlsx
2. Trouve les 8 alumni par leur nom de famille
3. Ajoute le texte vedette dans leur colonne Notes
4. Sauvegarde

⚠️ Ferme Excel avant de lancer !
"""

from openpyxl import load_workbook
from pathlib import Path

# ============================================================================
# CONFIG
# ============================================================================

EXCEL_FILE = "alumni-gaea21-AG-READY.xlsx"
SHEET_NAME = "Alumni AG"

# Les 8 vedettes avec leur texte à insérer (identifiés par nom de famille)
VEDETTES = {
    "BELLINI": "⭐⭐⭐⭐ VEDETTE AG MAX - Honda Motor Europe (logo majeur) - LCA certifiée - 10+ ans exp - a construit sustainability toolbox Gaea21 + app Répertoire Vert - Genève - TÉMOIGNAGE PRIORITÉ MAXIMALE",
    "BACHOU": "⭐⭐⭐ VEDETTE AG - Maroc 🇲🇦 - jeune diplômé INSEA - Stage Bank Al-Maghrib - finance durable + empreinte carbone - 1132 abonnés LinkedIn - TÉMOIGNAGE IDÉAL",
    "ASSENGHOUR": "⭐⭐⭐ VEDETTE AG - PhD agro-écologie Allemagne - parcours Maroc→Gaea21→Allemagne - recherche agriculture durable - TÉMOIGNAGE IDÉAL",
    "ACUÑA ANDREY": "⭐⭐⭐ VEDETTE AG - Mexique 🇲🇽 - Agroécologie communautés mayas - ex-coordinatrice Sustainability Club Gaea21 - TÉMOIGNAGE IDÉAL parcours international",
    "AMORY": "⭐⭐⭐ VEDETTE AG - Bilan carbone Numtech - Fresque Climat - parcours exemplaire Gaea21 → décarbonation entreprises - TÉMOIGNAGE IDÉAL",
    "AL NASIR": "⭐⭐ VEDETTE AG - Agriculture urbaine vertical farming - Wesh Grow Paris - parcours exemplaire - témoignage potentiel",
    "AGORO": "⭐ VEDETTE AG - Data Scientist - étudiant-entrepreneur SMA - ProNoïa Pau - parcours entrepreneurial",
    "BASSIL": "⭐⭐ VEDETTE AG - SKEMA Business School - Finance Project Manager Gaea21 - parcours international Liban/Dubai/France",
}


def main():
    # Vérifie que le fichier existe
    if not Path(EXCEL_FILE).exists():
        print(f"❌ Fichier introuvable : {EXCEL_FILE}")
        print("   Assure-toi d'être dans le dossier alumni_workflow")
        return

    print(f"📂 Ouverture de {EXCEL_FILE}...")

    try:
        wb = load_workbook(EXCEL_FILE)
    except PermissionError:
        print(f"❌ Impossible d'ouvrir {EXCEL_FILE}")
        print("   → Ferme Excel/Numbers d'abord !")
        return

    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ Feuille '{SHEET_NAME}' introuvable")
        print(f"   Feuilles disponibles : {wb.sheetnames}")
        return

    ws = wb[SHEET_NAME]

    # Trouve la ligne d'en-tête et les colonnes Nom et Notes
    header_row = None
    col_nom = None
    col_notes = None

    for row_idx in [1, 2, 3]:
        try:
            headers = [str(ws.cell(row=row_idx, column=c).value or "").strip()
                       for c in range(1, 30)]
            # Cherche "Nom" et "Notes"
            for idx, h in enumerate(headers, start=1):
                if h == "Nom":
                    col_nom = idx
                elif h == "Notes":
                    col_notes = idx
            if col_nom and col_notes:
                header_row = row_idx
                break
        except Exception:
            continue

    if not col_nom or not col_notes:
        print("❌ Colonnes 'Nom' ou 'Notes' introuvables")
        return

    print(f"✅ Ligne d'en-tête : {header_row}")
    print(f"✅ Colonne Nom : {col_nom} | Colonne Notes : {col_notes}")
    print()

    # Parcourt les lignes et marque les vedettes
    data_start = header_row + 1
    found = {}

    for row_idx in range(data_start, ws.max_row + 1):
        nom_cellule = ws.cell(row=row_idx, column=col_nom).value
        if not nom_cellule:
            continue

        nom_str = str(nom_cellule).strip().upper()

        # Cherche correspondance avec une vedette
        for vedette_nom, note_text in VEDETTES.items():
            if vedette_nom in nom_str:
                # Ajoute/Remplace la note
                existing_note = ws.cell(row=row_idx, column=col_notes).value or ""
                existing_str = str(existing_note).strip()

                # Si la note contient déjà "VEDETTE", on ne la dédouble pas
                if "VEDETTE" in existing_str.upper():
                    print(f"   ⏩ {vedette_nom} : déjà marquée vedette, skip")
                    found[vedette_nom] = "déjà marquée"
                else:
                    # Ajoute la note vedette (combine avec note existante si présente)
                    if existing_str and existing_str.lower() != "nan":
                        new_note = f"{note_text} | {existing_str}"
                    else:
                        new_note = note_text
                    ws.cell(row=row_idx, column=col_notes, value=new_note)
                    print(f"   ✨ {vedette_nom} : marqué vedette (ligne {row_idx})")
                    found[vedette_nom] = "ajouté"
                break

    # Résumé
    print()
    print(f"📊 Résumé : {len(found)}/8 vedettes trouvées")

    missing = [v for v in VEDETTES if v not in found]
    if missing:
        print(f"⚠️ Non trouvés dans le fichier : {missing}")

    # Sauvegarde
    try:
        wb.save(EXCEL_FILE)
        print(f"✅ Fichier sauvegardé : {EXCEL_FILE}")
        print()
        print("🎉 Terminé ! Relance ton app Streamlit pour voir les vedettes dans la page Brief.")
    except PermissionError:
        print("❌ Impossible de sauvegarder. Ferme Excel et relance.")


if __name__ == "__main__":
    main()
